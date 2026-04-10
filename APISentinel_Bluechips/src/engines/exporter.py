"""
Result Exporter for API Security Scanner
"""

import json
import csv
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from src.core.models import TestResult, SensitiveInfo


class Exporter:
    """Result exporter for API Security Scanner"""
    
    def __init__(self):
        """Initialize exporter"""
        pass
    
    def export_to_excel(self, results: List[TestResult], file_path: str) -> str:
        """Export results to Excel file"""
        wb = Workbook()
        
        # Create statistics sheet
        stats_sheet = wb.active
        stats_sheet.title = "Statistics"
        self._generate_statistics(stats_sheet, results)
        
        # Create results sheet
        results_sheet = wb.create_sheet(title="Results")
        self._generate_results_sheet(results_sheet, results)
        
        # Save workbook
        wb.save(file_path)
        return file_path
    
    def _generate_statistics(self, sheet, results: List[TestResult]):
        """Generate statistics sheet"""
        # Header
        sheet['A1'] = "API Security Scanner - Statistics"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:B1')
        
        # Statistics data
        total_requests = len(results)
        success_count = sum(1 for r in results if 200 <= r.response_status < 300)
        error_count = sum(1 for r in results if r.response_status >= 400)
        sensitive_count = sum(len(r.sensitive_info) for r in results)
        avg_response_time = sum(r.response_time for r in results) / total_requests if total_requests > 0 else 0
        
        stats = [
            ["Total Requests:", total_requests],
            ["Successful Requests (200-299):", success_count],
            ["Error Requests (400+):", error_count],
            ["Total Sensitive Information Found:", sensitive_count],
            ["Average Response Time (s):", f"{avg_response_time:.2f}"],
            ["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for i, (label, value) in enumerate(stats, start=3):
            sheet[f'A{i}'] = label
            sheet[f'B{i}'] = value
            sheet[f'A{i}'].font = Font(bold=True)
        
        # Auto adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
    
    def _generate_results_sheet(self, sheet, results: List[TestResult]):
        """Generate results sheet"""
        # Headers
        headers = [
            "ID", "Method", "URL", "Status", "Length", "Time (s)",
            "Sensitive Count", "Sensitive Details", "Request Headers",
            "Request Body", "Response Headers", "Response Body"
        ]
        
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
            sheet.cell(row=1, column=col).font = Font(bold=True)
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill data
        for row, result in enumerate(results, start=2):
            # Basic info
            sheet.cell(row=row, column=1, value=row-1)
            sheet.cell(row=row, column=2, value=result.endpoint.method)
            sheet.cell(row=row, column=3, value=result.endpoint.url)
            
            # Status code with color
            status_cell = sheet.cell(row=row, column=4, value=result.response_status)
            if 200 <= result.response_status < 300:
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif 400 <= result.response_status < 500:
                status_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light orange
            elif result.response_status >= 500:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
            
            # Other data
            sheet.cell(row=row, column=5, value=result.response_length)
            sheet.cell(row=row, column=6, value=f"{result.response_time:.2f}")
            
            # Sensitive info
            sensitive_count = len(result.sensitive_info)
            sheet.cell(row=row, column=7, value=sensitive_count)
            
            # Sensitive details
            sensitive_details = "\n".join([
                f"{info.rule_name}: {info.matched_content}" 
                for info in result.sensitive_info
            ])
            sheet.cell(row=row, column=8, value=sensitive_details)
            
            # Request headers
            request_headers = "\n".join([f"{k}: {v}" for k, v in result.request_headers.items()])
            sheet.cell(row=row, column=9, value=request_headers)
            
            # Request body
            sheet.cell(row=row, column=10, value=result.request_body)
            
            # Response headers
            response_headers = "\n".join([f"{k}: {v}" for k, v in result.response_headers.items()])
            sheet.cell(row=row, column=11, value=response_headers)
            
            # Response body
            sheet.cell(row=row, column=12, value=result.response_body)
        
        # Auto adjust column widths
        column_widths = [10, 10, 30, 10, 10, 10, 15, 30, 30, 30, 30, 50]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width
    
    def export_to_csv(self, results: List[TestResult], file_path: str) -> str:
        """Export results to CSV file"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow([
                "ID", "Method", "URL", "Status", "Length", "Time (s)",
                "Sensitive Count", "Sensitive Details", "Request Headers",
                "Request Body", "Response Headers", "Response Body"
            ])
            
            # Data
            for i, result in enumerate(results, start=1):
                sensitive_details = "\n".join([
                    f"{info.rule_name}: {info.matched_content}" 
                    for info in result.sensitive_info
                ])
                
                request_headers = "\n".join([f"{k}: {v}" for k, v in result.request_headers.items()])
                response_headers = "\n".join([f"{k}: {v}" for k, v in result.response_headers.items()])
                
                writer.writerow([
                    i,
                    result.endpoint.method,
                    result.endpoint.url,
                    result.response_status,
                    result.response_length,
                    f"{result.response_time:.2f}",
                    len(result.sensitive_info),
                    sensitive_details,
                    request_headers,
                    result.request_body,
                    response_headers,
                    result.response_body
                ])
        
        return file_path
    
    def export_to_json(self, results: List[TestResult], file_path: str) -> str:
        """Export results to JSON file"""
        export_data = {
            "metadata": {
                "export_date": datetime.now().isoformat(),
                "total_results": len(results),
                "success_count": sum(1 for r in results if 200 <= r.response_status < 300),
                "error_count": sum(1 for r in results if r.response_status >= 400),
                "sensitive_count": sum(len(r.sensitive_info) for r in results)
            },
            "results": []
        }
        
        for i, result in enumerate(results, start=1):
            result_data = {
                "id": i,
                "method": result.endpoint.method,
                "url": result.endpoint.url,
                "status": result.response_status,
                "length": result.response_length,
                "time": result.response_time,
                "sensitive_count": len(result.sensitive_info),
                "sensitive_info": [
                    {
                        "rule_name": info.rule_name,
                        "rule_level": info.rule_level,
                        "matched_content": info.matched_content
                    }
                    for info in result.sensitive_info
                ],
                "request": {
                    "headers": result.request_headers,
                    "body": result.request_body
                },
                "response": {
                    "headers": result.response_headers,
                    "body": result.response_body
                }
            }
            export_data["results"].append(result_data)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        return file_path
    
    def export_to_html(self, results: List[TestResult], file_path: str) -> str:
        """Export results to HTML file"""
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>API Security Scanner Results</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        h2 {{ color: #666; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        tr:hover {{ background-color: #f5f5f5; }}
        .status-200 {{ background-color: #90EE90; }}
        .status-400 {{ background-color: #FFCC99; }}
        .status-500 {{ background-color: #FFB6C1; }}
        .stats {{ background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        .stats-item {{ display: inline-block; margin-right: 20px; }}
        .stats-label {{ font-weight: bold; }}
    </style>
</head>
<body>
    <h1>API Security Scanner Results</h1>
    <div class="stats">
        <h2>Statistics</h2>
        <div class="stats-item"><span class="stats-label">Total Requests:</span> {total_requests}</div>
        <div class="stats-item"><span class="stats-label">Successful:</span> {success_count}</div>
        <div class="stats-item"><span class="stats-label">Errors:</span> {error_count}</div>
        <div class="stats-item"><span class="stats-label">Sensitive Info:</span> {sensitive_count}</div>
        <div class="stats-item"><span class="stats-label">Export Date:</span> {export_date}</div>
    </div>
    <h2>Detailed Results</h2>
    <table>
        <tr>
            <th>ID</th>
            <th>Method</th>
            <th>URL</th>
            <th>Status</th>
            <th>Length</th>
            <th>Time (s)</th>
            <th>Sensitive Count</th>
            <th>Sensitive Details</th>
        </tr>
        {table_rows}
    </table>
</body>
</html>
"""
        
        # Calculate statistics
        total_requests = len(results)
        success_count = sum(1 for r in results if 200 <= r.response_status < 300)
        error_count = sum(1 for r in results if r.response_status >= 400)
        sensitive_count = sum(len(r.sensitive_info) for r in results)
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Generate table rows
        table_rows = []
        for i, result in enumerate(results, start=1):
            # Determine status class
            if 200 <= result.response_status < 300:
                status_class = "status-200"
            elif 400 <= result.response_status < 500:
                status_class = "status-400"
            else:
                status_class = "status-500"
            
            # Sensitive details
            sensitive_details = "<br>".join([
                f"{info.rule_name}: {info.matched_content}" 
                for info in result.sensitive_info
            ]) if result.sensitive_info else "None"
            
            row = f"""
        <tr>
            <td>{i}</td>
            <td>{result.endpoint.method}</td>
            <td>{result.endpoint.url}</td>
            <td class="{status_class}">{result.response_status}</td>
            <td>{result.response_length}</td>
            <td>{result.response_time:.2f}</td>
            <td>{len(result.sensitive_info)}</td>
            <td>{sensitive_details}</td>
        </tr>
            """
            table_rows.append(row)
        
        # Fill template
        html_content = html_content.format(
            total_requests=total_requests,
            success_count=success_count,
            error_count=error_count,
            sensitive_count=sensitive_count,
            export_date=export_date,
            table_rows="".join(table_rows)
        )
        
        # Write to file
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return file_path
